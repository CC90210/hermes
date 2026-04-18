from __future__ import annotations

import io
import json
import os
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import httpx
from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parents[1] / ".env")

OLLAMA_HOST: str = os.environ.get("OLLAMA_HOST", "http://localhost:11434")
OLLAMA_MODEL: str = os.environ.get("OLLAMA_MODEL", "qwen2.5:32b")

_EXTRACTION_PROMPT = """\
You are a structured data extraction engine for wholesale purchase orders.
Extract every field from the text below and return ONLY a valid JSON object.
Do not include markdown fences, explanations, or any text outside the JSON.

Required JSON shape:
{
  "po_number": "<string or null>",
  "customer_name": "<string or null>",
  "customer_email": "<string or null>",
  "customer_address": "<string or null>",
  "ship_to_address": "<string or null>",
  "order_date": "<YYYY-MM-DD or null>",
  "ship_date": "<YYYY-MM-DD or null>",
  "notes": "<string or null>",
  "line_items": [
    {
      "sku": "<string or null>",
      "description": "<string>",
      "quantity": <integer>,
      "unit_price": <float>,
      "upc": "<string or null>"
    }
  ]
}

Rules:
- po_number: vendor/buyer PO reference number (e.g. "PO-12345", "4500123456")
- customer_name: the BUYER (bill-to) company or person name
- customer_address: bill-to / sold-to address as a single string
- ship_to_address: delivery / ship-to address; copy customer_address if identical
- order_date: the date the PO was issued
- ship_date: requested delivery or ship date
- line_items: every product line — include ALL lines, never skip
- sku: vendor or buyer item code; use whatever code appears in the document
- quantity: always a positive integer; infer from "EA", "CS", "EACH" etc.
- unit_price: per-unit price as float; 0.0 if not present
- upc: 12-digit UPC/EAN if present, else null
- notes: any special instructions, shipping terms, or comments

Purchase Order text:
---
{text}
---
"""


@dataclass
class LineItem:
    sku: Optional[str]
    description: str
    quantity: int
    unit_price: float
    upc: Optional[str] = None


@dataclass
class POData:
    po_number: Optional[str]
    customer_name: Optional[str]
    customer_email: Optional[str]
    customer_address: Optional[str]
    ship_to_address: Optional[str]
    order_date: Optional[str]
    ship_date: Optional[str]
    line_items: list[LineItem] = field(default_factory=list)
    notes: Optional[str] = None
    raw_text: str = ""
    # Set by the storage layer after the order is persisted
    internal_order_id: Optional[str] = None

    @classmethod
    def from_db_row(cls, row: dict[str, Any]) -> POData:
        """Reconstruct a POData instance from a flattened orders DB row.

        Line items are not stored on the orders table; callers that need them
        should load them separately via storage.db.get_order_lines().
        """
        return cls(
            po_number=row.get("po_number"),
            customer_name=row.get("customer_name"),
            customer_email=row.get("customer_email"),
            customer_address=row.get("customer_address"),
            ship_to_address=row.get("ship_to_address"),
            order_date=row.get("order_date"),
            ship_date=row.get("ship_date"),
            notes=row.get("notes"),
            raw_text=row.get("raw_text", ""),
            internal_order_id=str(row["id"]) if row.get("id") is not None else None,
        )


# ---------------------------------------------------------------------------
# Format extractors
# ---------------------------------------------------------------------------

def _extract_text_pdf(content: bytes) -> str:
    import pdfplumber  # lazy import — optional dep

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        pages = [page.extract_text() or "" for page in pdf.pages]
    return "\n".join(pages)


def _extract_text_excel(content: bytes) -> str:
    from openpyxl import load_workbook  # lazy import

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    rows: list[str] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append("\t".join(cells))
    wb.close()
    return "\n".join(rows)


def _extract_text_html(content: bytes) -> str:
    text = content.decode("utf-8", errors="replace")
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _extract_text_edi(content: bytes) -> str:
    """Minimal X12 850 segment parser.

    Extracts BEG (PO number / date), N1 (names), PO1 (line items), and CTT
    (totals), then returns a human-readable summary for the LLM to parse.
    """
    raw = content.decode("utf-8", errors="replace")

    # Detect element and segment separators from ISA envelope
    element_sep = raw[3] if len(raw) > 3 else "*"
    segment_sep = raw[105] if len(raw) > 105 else "~"

    segments = [s.strip() for s in raw.split(segment_sep) if s.strip()]
    lines: list[str] = []
    current_party: Optional[str] = None

    for seg in segments:
        elements = seg.split(element_sep)
        tag = elements[0].strip()

        if tag == "BEG":
            po_num = elements[3] if len(elements) > 3 else ""
            po_date_raw = elements[5] if len(elements) > 5 else ""
            po_date = (
                f"{po_date_raw[:4]}-{po_date_raw[4:6]}-{po_date_raw[6:8]}"
                if len(po_date_raw) == 8
                else po_date_raw
            )
            lines.append(f"PO Number: {po_num}")
            lines.append(f"Order Date: {po_date}")

        elif tag == "N1":
            qualifier = elements[1] if len(elements) > 1 else ""
            name = elements[2] if len(elements) > 2 else ""
            label = {"BT": "Bill-To", "ST": "Ship-To", "BY": "Buyer", "SE": "Seller"}.get(
                qualifier, qualifier
            )
            current_party = label
            lines.append(f"{label} Name: {name}")

        elif tag in ("N3", "N4") and current_party:
            addr = element_sep.join(elements[1:])
            lines.append(f"{current_party} Address: {addr}")

        elif tag == "PO1":
            qty_raw = elements[2] if len(elements) > 2 else "0"
            price_raw = elements[4] if len(elements) > 4 else "0"
            qualifiers: dict[str, str] = {}
            idx = 6
            while idx < len(elements) - 1:
                qualifiers[elements[idx]] = elements[idx + 1]
                idx += 2
            upc = qualifiers.get("UI", "")
            sku = qualifiers.get("BP", qualifiers.get("VN", qualifiers.get("IN", "")))
            desc = qualifiers.get("PD", "")
            try:
                qty = int(float(qty_raw))
            except ValueError:
                qty = 0
            try:
                price = float(price_raw)
            except ValueError:
                price = 0.0
            lines.append(
                f"Line Item: SKU={sku} UPC={upc} Qty={qty} UnitPrice={price} Desc={desc}"
            )

        elif tag == "CTT":
            total_lines = elements[1] if len(elements) > 1 else ""
            lines.append(f"Total Line Count: {total_lines}")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# LLM extraction
# ---------------------------------------------------------------------------

def _call_ollama(text: str) -> dict[str, Any]:
    prompt = _EXTRACTION_PROMPT.replace("{text}", text[:12000])
    payload = {
        "model": OLLAMA_MODEL,
        "prompt": prompt,
        "stream": False,
        "format": "json",
        "options": {"temperature": 0.0},
    }
    resp = httpx.post(
        f"{OLLAMA_HOST}/api/generate",
        json=payload,
        timeout=120.0,
    )
    resp.raise_for_status()
    raw_response = resp.json().get("response", "{}")
    return json.loads(raw_response)


def _build_po_data(extracted: dict[str, Any], raw_text: str) -> POData:
    raw_items: list[dict[str, Any]] = extracted.get("line_items") or []
    line_items: list[LineItem] = []
    for item in raw_items:
        try:
            qty = int(item.get("quantity") or 0)
        except (TypeError, ValueError):
            qty = 0
        try:
            price = float(item.get("unit_price") or 0.0)
        except (TypeError, ValueError):
            price = 0.0
        line_items.append(
            LineItem(
                sku=item.get("sku") or None,
                description=str(item.get("description") or ""),
                quantity=qty,
                unit_price=price,
                upc=item.get("upc") or None,
            )
        )
    return POData(
        po_number=extracted.get("po_number") or None,
        customer_name=extracted.get("customer_name") or None,
        customer_email=extracted.get("customer_email") or None,
        customer_address=extracted.get("customer_address") or None,
        ship_to_address=extracted.get("ship_to_address") or None,
        order_date=extracted.get("order_date") or None,
        ship_date=extracted.get("ship_date") or None,
        line_items=line_items,
        notes=extracted.get("notes") or None,
        raw_text=raw_text,
    )


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_po(content: bytes, filename: str, content_type: str) -> POData:
    """Parse a purchase order from raw bytes.

    Detects format from the filename extension and content_type, extracts
    plain text, then sends it to the local Ollama model for structured
    extraction.
    """
    name_lower = filename.lower()
    ct_lower = content_type.lower()

    if name_lower.endswith(".pdf") or "pdf" in ct_lower:
        raw_text = _extract_text_pdf(content)
    elif name_lower.endswith((".xlsx", ".xls")) or "spreadsheet" in ct_lower or "excel" in ct_lower:
        raw_text = _extract_text_excel(content)
    elif name_lower.endswith((".edi", ".x12", ".850")) or "edi" in ct_lower:
        raw_text = _extract_text_edi(content)
    elif "html" in ct_lower or name_lower.endswith((".html", ".htm")):
        raw_text = _extract_text_html(content)
    else:
        # Plain text, email body, CSV, etc.
        raw_text = content.decode("utf-8", errors="replace")

    extracted = _call_ollama(raw_text)
    return _build_po_data(extracted, raw_text)


class POParser:
    """Stateful wrapper around the functional ``parse_po`` / ``validate_po`` helpers.

    The orchestrator instantiates one ``POParser`` at startup and calls
    ``parse_and_persist`` for every inbound email dict.
    """

    def __init__(self, config: Any) -> None:  # config is manager.config.AppConfig
        self._config = config

    async def parse_and_persist(self, raw_email: dict[str, Any]) -> int:
        """Parse a raw email dict into structured PO data and persist to the DB.

        raw_email shape::

            {
                "uid": str,
                "subject": str,
                "from": str,
                "attachments": [(filename, content_type, bytes), ...],
                "body": str,
            }

        Returns the ``order_id`` (int) of the newly created orders row.
        """
        from storage.db import create_order, create_order_line, log_audit, OrderStatus

        attachments: list[tuple[str, str, bytes]] = raw_email.get("attachments") or []

        po: POData | None = None
        for filename, content_type, content in attachments:
            po = parse_po(content, filename, content_type)
            break  # use the first PO-looking attachment

        if po is None:
            # Fall back to the email body
            body_bytes = (raw_email.get("body") or "").encode()
            po = parse_po(body_bytes, "email_body.txt", "text/plain")

        db_path = self._config.db_path
        order_id = await create_order(
            db_path,
            po_number=po.po_number or "UNKNOWN",
            customer_name=po.customer_name or "UNKNOWN",
            customer_email=po.customer_email or raw_email.get("from", ""),
            raw_email_id=str(raw_email.get("uid", "")),
            status=OrderStatus.PARSED,
        )

        for item in po.line_items:
            await create_order_line(
                db_path,
                order_id=order_id,
                sku=item.sku or "UNKNOWN",
                description=item.description,
                quantity=float(item.quantity),
                unit_price=item.unit_price,
                line_total=item.quantity * item.unit_price,
            )

        await log_audit(
            db_path,
            agent_name="po_parser",
            action="order_parsed",
            details={"order_id": order_id, "po_number": po.po_number, "uid": raw_email.get("uid")},
        )
        return order_id


def validate_po(po: POData) -> list[str]:
    """Return a list of human-readable validation errors. Empty list = valid."""
    errors: list[str] = []

    if not po.po_number or not po.po_number.strip():
        errors.append("Missing PO number.")

    if not po.customer_name or not po.customer_name.strip():
        errors.append("Missing customer name.")

    if not po.line_items:
        errors.append("No line items found.")
    else:
        for i, item in enumerate(po.line_items, start=1):
            if item.quantity <= 0:
                errors.append(
                    f"Line {i}: quantity must be a positive integer (got {item.quantity})."
                )
            if not item.description.strip():
                errors.append(f"Line {i}: missing description.")

    return errors
